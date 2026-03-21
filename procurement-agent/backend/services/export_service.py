""" CSV/Excel Export Service """

import csv
import io
from datetime import datetime
from typing import Literal

def generate_csv(quotes: list[dict], prodect_name: str) -> bytes:
    output = io.StringIO()

    fieldnames = [
        "Rank",
        "Supplier Name",
        "website",
        "Unit Price",
        "Total Price",
        "Delivery Days",
        "Min Order Qty",
        "Payment Terms",
        "Score",
        "Recommended",
        "Notes",
        "Fetched At"
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
 
    # Header info rows
    output.write(f"Product: {product_name}\n")
    output.write(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}\n")
    output.write(f"Total Quotes: {len(quotes)}\n\n")
 
    writer.writeheader()
 
    for i, q in enumerate(quotes, 1):
        writer.writerow({
            "Rank":               i,
            "Supplier Name":      q.get("supplier_name", ""),
            "Website":            q.get("supplier_url", ""),
            "Unit Price (INR)":   q.get("unit_price", ""),
            "Total Price (INR)":  q.get("total_price", ""),
            "Delivery Days":      q.get("delivery_days", ""),
            "Min Order Qty":      q.get("minimum_order_qty", ""),
            "Payment Terms":      q.get("payment_terms", ""),
            "Score":              round(q.get("score", 0), 1) if q.get("score") else "",
            "Recommended":        "YES" if q.get("is_recommended") else "",
            "Notes":              q.get("additional_notes", ""),
            "Fetched At":         q.get("fetched_at", "")
        })
 
    return output.getvalue().encode("utf-8")
 
 
def generate_excel(quotes: list[dict], product_name: str) -> bytes:
    """
    Generate Excel bytes using openpyxl.
    Falls back to CSV if openpyxl not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
 
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quotes"
 
        # Title row
        ws["A1"] = f"Procurement Quotes — {product_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws["A3"] = f"Total Quotes: {len(quotes)}"
 
        # Header row
        headers = [
            "Rank", "Supplier", "Website", "Unit Price",
            "Total Price", "Delivery Days", "Min Order",
            "Payment Terms", "Score", "Recommended", "Notes"
        ]
 
        header_row = 5
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1D4ED8")
            cell.alignment = Alignment(horizontal="center")
 
        # Data rows
        for i, q in enumerate(quotes, 1):
            row = header_row + i
            ws.cell(row=row, column=1,  value=i)
            ws.cell(row=row, column=2,  value=q.get("supplier_name", ""))
            ws.cell(row=row, column=3,  value=q.get("supplier_url", ""))
            ws.cell(row=row, column=4,  value=q.get("unit_price"))
            ws.cell(row=row, column=5,  value=q.get("total_price"))
            ws.cell(row=row, column=6,  value=q.get("delivery_days", ""))
            ws.cell(row=row, column=7,  value=q.get("minimum_order_qty", ""))
            ws.cell(row=row, column=8,  value=q.get("payment_terms", ""))
            ws.cell(row=row, column=9,  value=round(q.get("score", 0), 1) if q.get("score") else "")
            ws.cell(row=row, column=10, value="✅ YES" if q.get("is_recommended") else "")
            ws.cell(row=row, column=11, value=q.get("additional_notes", ""))
 
            # Highlight recommended row in green
            if q.get("is_recommended"):
                for col in range(1, 12):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="D1FAE5")
 
        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
 
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
 
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return generate_csv(quotes, product_name)